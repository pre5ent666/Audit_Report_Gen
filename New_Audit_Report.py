from openpyxl import load_workbook, workbook
from openpyxl import Workbook
import re
import glob
import os

#--------------------------------------------- Input Settings -------------------------------------------------
tail = 2
Week = 46
cmp = Week - 1
order1 = ['Main', 'STR', 'SORP']    # This week
order2 = ['Main', 'STR', 'SORP']    # Last week
output_order = ['Main', 'STR', 'SORP']
folder = os.getcwd()
#--------------------------------------------------------------------------------------------------------------

def All_Results(Week, folder, Order):
    os.mkdir(folder + '\\All')
    output = Workbook() # Result Summary
    for Line_Name in Order:
        print('# ' + Line_Name + ' #')
        Title = 'W' + str(Week) + '_' + Line_Name
        ws = output.create_sheet(Line_Name) # create sheet for each line
        row_title = ['Original GM TC ID', Title] 
        ws.append(row_title) # add title

        files = glob.glob(folder + Title +'_**.xlsx', recursive = True)
        for xls in files :
            wb =  load_workbook(xls)
            Results_Sum(wb, ws, xls)
    output.save(folder + '\\All\\2021_W' + str(Week) + '.xlsx')

def Results_Sum(wb, ws, path):
    j = 0
    for sheet in wb:
        i = 0
        if sheet.title == 'Case need update'or sheet.title == 'Summary' or sheet.title == 'summary' or sheet.title == 'Sheet':
            break
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[0] is None:
                ws.append(row)
                i = i + 1
                j = j + 1
        print(sheet.title + ': ' + str(i))
    print('====================================================================================='\
     + '\nFrom ' + path\
     + '\nTotal:' + str(j)\
     + '\n=====================================================================================\n')

def matching(wb1, wb2, wb_out, tail, order1, order2, output_order):
    # wb_out = Workbook()
    for j in range(len(order1)):
        sheet1 = wb1[order1[j]]
        sheet2 = wb2[order2[j]]
        ws = wb_out.create_sheet(output_order[j])
        i = 0
        for row1 in sheet1.iter_rows(max_col=tail, values_only=True):
            cell_output = cell_data(row1)
            match = False
            if cell_output[0] == 'none':
                break
            for row2 in sheet2.iter_rows(max_col=2, values_only=True):
                cell_temp = cell_data(row2)
                if str(cell_output[0]).lower() == str(cell_temp[0]).lower():
                    match = True
                    cell_output.append(cell_temp[1])
                    break
            if match == False:
                cell_output.append('none')
            i = i + 1
            cell_output.append('=EQ(Left($B'+ str(i) + ',4),Left($C' + str(i) + ',4))')
            ws.append(cell_output)
        ws["D1"] = 'Same'
        print(output_order[j] + ': ' + str(i-1))

def cell_data(row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

# -------------------------------- Results Summary -----------------------------------
All_Results(Week, folder, order1)
All_Results(Week, folder, order2)
# -------------------------------- Matching case results -------------------------------
wb1 = load_workbook(folder + '\\All\\2021_W' +  str(Week)  + '.xlsx')
wb2 = load_workbook(folder + '\\All\\2021_W' + str(cmp) + '.xlsx')
wb_out = Workbook()
os.mkdir(folder + '\\Audit_Report')
matching(wb1, wb2, wb_out, tail, order1, order2, output_order)
wb_out.save(folder + '\\Audit_Report\\Audit_Report_W' + str(Week) + '.xlsx')