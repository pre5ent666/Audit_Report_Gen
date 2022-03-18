from openpyxl import load_workbook
from openpyxl import Workbook
import re

#--------------------------------------------- Input Settings -------------------------------------------------
tail = 7
Week = 202210
Cmp = 202209
order1 = ['SORP_Resume', 'Main_Cold']    # This week
order2 = ['SORP_Disable', 'Main_Resume']    # Last week
output_order = order1

folder = 'C:\\Users\\Yvonne\\Documents\\Results'
#--------------------------------------------------------------------------------------------------------------

def matching(wb1, wb2, wb_out, tail, order1, order2, output_order):
    # wb_out = Workbook()
    for j in range(len(order1)):
        sheet1 = wb1[order1[j]]
        sheet2 = wb2[order2[j]]
        ws = wb_out.create_sheet(output_order[j])
        i = 0
        for row1 in sheet1.iter_rows(max_col=tail, values_only=True):
            cell_temp1 = cell_data(row1)
            cell_output = cell_temp1[0:2]
            match = False
            if cell_output[0] == 'none':
                break
            for row2 in sheet2.iter_rows(max_col=7, values_only=True):
                cell_temp2 = cell_data(row2)
                if str(cell_output[0]).lower() == str(cell_temp2[0]).lower():
                    match = True
                    cell_output.append(cell_temp2[1])
                    break
            if match == False:
                cell_output.append('none')
            i = i + 1
            cell_output.append('=EXACT(Left($B'+ str(i) + ',4),Left($C' + str(i) + ',4))')
            cell_output.append(cell_temp1[6])
            cell_output.append(cell_temp2[6])
            ws.append(cell_output)
        ws["D1"] = 'Same'
        print(output_order[j] + ': ' + str(i-1))

def cell_data(row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

wb1 = load_workbook(folder + '\\All\\'+ str(Week)[:4]+ '_W' + (str(Week)[4:]).zfill(2) + '.xlsx')
wb2 = load_workbook(folder + '\\All\\'+ str(Cmp)[:4]+ '_W' + (str(Cmp)[4:]).zfill(2) + '.xlsx')
print(folder + '\\All\\'+ str(Cmp)[:4]+ '_W' + (str(Cmp)[4:]).zfill(2) + '.xlsx')
wb_out = Workbook()
matching(wb1, wb2, wb_out, tail, order1, order2, output_order)
wb_out.save(folder + '\\Audit_Report\\Audit_Report_W' + (str(Week)[4:]).zfill(2) + '.xlsx')
