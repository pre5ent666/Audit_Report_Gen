from openpyxl import load_workbook, workbook
from openpyxl import Workbook
import re
import glob
import pathlib

#--------------------------------------------- Input Settings -------------------------------------------------
Week1 = 202235
Week2 = 202234
order1 = ['Main_Resume']    # Line_Mode of Week1
order2 = ['Main_Resume']    # Line_Mode of Week2
output_order = order1       # sheet name of Audit Report
Enable_Productlines_Compare = False

folder = str(pathlib.Path(__file__).parent.resolve())
tail = 4
#--------------------------------------------------------------------------------------------------------------

def Results_Sum(wb, ws, path):
    j = 0
    for sheet in wb:
        i = 0
        if sheet.title == 'Case need update'or sheet.title == 'Summary' or sheet.title == 'summary' or sheet.title == 'Sheet':
            break
        for row in sheet.iter_rows(min_row=2, max_col=tail, values_only=True):
            if not row[0] is None:
                ws.append(row)
                i = i + 1
                j = j + 1
        print(sheet.title + ': ' + str(i))
    print('====================================================================================='\
     + '\nFrom ' + path\
     + '\nTotal:' + str(j)\
     + '\n=====================================================================================\n')

def Results_All(Order,Week1):
    output = Workbook() # Result Summary
    for Line_Name in Order:
        print('# ' + Line_Name + ' #')
        Title = 'W' + (str(Week1)[4:]).zfill(2) + '_' + Line_Name
        ws = output.create_sheet(Line_Name, 0) # create sheet for each line
        row_title = ['Original GM TC ID', Title] 
        ws.append(row_title) # add title

        files = glob.glob(folder + '\\Sorted\\' + Title +'**.xlsx', recursive = True)
        for xls in files :
            wb =  load_workbook(xls)
            Results_Sum(wb, ws, xls)
    output.save(folder + '\\All\\' + str(Week1)[:4]+ '_W' + (str(Week1)[4:]).zfill(2) + '.xlsx')

def matching(sheet1, sheet2, sheet_out, tail):
    i = 0
    for row1 in sheet1.iter_rows(max_col=tail, values_only=True):
        cell_temp1 = cell_data(row1)
        cell_output = cell_temp1[0:2]
        match = False
        if cell_output[0] == 'none':
            break
        for row2 in sheet2.iter_rows(max_col=tail, values_only=True):
            cell_temp2 = cell_data(row2)
            if str(cell_output[0]).lower() == str(cell_temp2[0]).lower():
                match = True
                cell_output.append(cell_temp2[1])
                break
        if match == False:
            cell_output.append('none')
            cell_temp2[0:3] = ['none','none','none','none']
        i = i + 1
        cell_output.append('=EXACT(Left($B'+ str(i) + ',4),Left($C' + str(i) + ',4))')
        cell_output.append(cell_temp1[3])
        cell_output.append(cell_temp2[3])
        sheet_out.append(cell_output)
    sheet_out["D1"] = 'Same Results'
    print(sheet_out.title + ': ' + str(i-1))

def cell_data(row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

# Put all sheets together
Results_All(order1,Week1)
Results_All(order2,Week2)
wb1 = load_workbook(folder + '\\All\\'+ str(Week1)[:4]+ '_W' + (str(Week1)[4:]).zfill(2) + '.xlsx')
wb2 = load_workbook(folder + '\\All\\'+ str(Week2)[:4]+ '_W' + (str(Week2)[4:]).zfill(2) + '.xlsx')

# Matching results
print("Matching results......")
wb_out = Workbook()
# - base on week
for j in range(len(order1)):
    matching(wb1[order1[j]], wb2[order2[j]], wb_out.create_sheet(output_order[j], 0), tail)
# - base on product line (for week1)
if Enable_Productlines_Compare:
    for k in range(len(order1)-1):
        matching(wb1[order1[k]], wb1[order1[k+1]], wb_out.create_sheet(order1[k]+"_"+order1[k+1], 0), tail)
wb_out.save(folder + '\\Audit_Report\\Audit_Report_W' + (str(Week1)[4:]).zfill(2) + '.xlsx')
