from openpyxl import load_workbook
from openpyxl import Workbook
import re

#--------------------------------------------- Input Settings -------------------------------------------------
tail = 7
Week = 202217
Cmp = 202216
order1 = ['SORP_Resume','Main_Resume']    # This week
order2 = ['SORP_Resume','Main_Cold']    # Last week
output_order = order1

folder = 'C:\\Users\\Yvonne\\Documents\\Results'
#--------------------------------------------------------------------------------------------------------------

def matching(sheet1, sheet2, sheet_out, tail):
    i = 0
    for row1 in sheet1.iter_rows(max_col=tail, values_only=True):
        cell_temp1 = cell_data(row1)
        cell_output = cell_temp1[0:2]
        match = False
        if cell_output[0] == 'none':
            break
        for row2 in sheet2.iter_rows(max_col=7, values_only=True):
            cell_temp2 = cell_data(row2)
            if str(cell_output[0]).lower() == str(cell_temp2[0]).lower():
                match = True
                cell_output.append(cell_temp2[1])
                break
        if match == False:
            cell_output.append('none')
        i = i + 1
        cell_output.append('=EXACT(Left($B'+ str(i) + ',4),Left($C' + str(i) + ',4))')
        cell_output.append(cell_temp1[6])
        cell_output.append(cell_temp2[6])
        sheet_out.append(cell_output)
    sheet_out["D1"] = 'Same Results'
    print(sheet_out.title + ': ' + str(i-1))

def cell_data(row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells


wb1 = load_workbook(folder + '\\All\\'+ str(Week)[:4]+ '_W' + (str(Week)[4:]).zfill(2) + '.xlsx')
wb2 = load_workbook(folder + '\\All\\'+ str(Cmp)[:4]+ '_W' + (str(Cmp)[4:]).zfill(2) + '.xlsx')

wb_out = Workbook()
for j in range(len(order1)):
    sheet1 = wb1[order1[j]]
    sheet2 = wb2[order2[j]]
    sheet_out = wb_out.create_sheet(output_order[j])
    matching(sheet1, sheet2, sheet_out, tail)
matching(wb1[order1[0]], wb1[order1[1]], wb_out.create_sheet("Weekly_lines"), tail)
wb_out.save(folder + '\\Audit_Report\\Audit_Report_W' + (str(Week)[4:]).zfill(2) + '.xlsx')
